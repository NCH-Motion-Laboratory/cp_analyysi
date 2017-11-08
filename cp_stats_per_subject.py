# -*- coding: utf-8 -*-
"""

CP-projektin analyysiskripti:
laske numeerista dataa, kaikki trialit / koehlöt

erikseen norm/kognitiivinen
erikseen oikea/vasen puoli


muuttujat:

hip flexion:
    strike, toeoff

knee flexion:
    strike
    maksimi ekan puoliskon ajalta (derivaatan nollakohta)

ankle dorsi/plant:
    strike, toeoff
    kontaktivaiheen aik. maksimi

thorax lateral flex:
    max, min

thorax rotation:
    max, min


@author: Jussi (jnu@iki.fi)
"""


from __future__ import print_function
import gaitutils
from gaitutils import cfg
import matplotlib.pyplot as plt
import glob
import numpy as np
import sys
import os
import os.path as op
import scipy


def _get_data(subject):
    """Average trials for a given subject, separately for cogn / normal"""

    # parameters
    rootdir = 'Z:\\Userdata_Vicon_Server\\CP-projekti'
    plotdir = "Z:\\CP_projekti_analyysit\\Normal_vs_cognitive"

    Cfiles = list()
    Nfiles = list()

    # try to auto find data dirs under subject dir
    subjdir = op.join(rootdir, subject)
    datadirs = [file for file in os.listdir(subjdir) if
                op.isdir(op.join(subjdir, file))]
    if len(datadirs) > 1:
        raise Exception('Multiple data dirs under subject')
    datadir = datadirs[0]

    # collect normal walk trials
    Nglob = op.join(subjdir, datadir, '*N?_*.c3d')
    Nfiles += glob.glob(Nglob)

    # collect cognitive trials
    Cglob = op.join(subjdir, datadir, '*C?_*.c3d')
    Cfiles += glob.glob(Cglob)

    if not (Cfiles and Nfiles):
        raise Exception('No trials for subject %s' % subject)

    print('%s: total of %d/%d trials' % (subject, len(Nfiles), len(Cfiles)))

    C_avgdata, C_stddata, _, _ = gaitutils.stats.average_trials(Cfiles)
    N_avgdata, N_stddata, _, _ = gaitutils.stats.average_trials(Nfiles)

    return N_avgdata, N_stddata, C_avgdata, C_stddata


def _get_vars(N_avgdata, C_avgdata):
    """Compute variables of interest from the averaged data"""

    results = dict()
    results['C'] = dict()
    results['N'] = dict()

    for cond in results.keys():
        for varname_ in stats_vars:
            for side in ['R', 'L']:
                varname = side + varname_
                results[cond][varname] = dict()
                data = C_avgdata if cond == 'C' else N_avgdata
                # data at first frame (foot strike)
                results[cond][varname]['at_foot_strike'] = data[varname][0]
                # maximum for each curve
                results[cond][varname]['max'] = data[varname].max()
                # minimum for each curve
                results[cond][varname]['min'] = data[varname].min()
                # local extrema for each curve
                xind = scipy.signal.argrelextrema(data[varname], np.greater)[0]
                results[cond][varname]['localextind'] = xind
                results[cond][varname]['localext'] = data[varname][xind]
                # local extrema during contact
                xind_contact = xind[np.where(xind < 60)]
                results[cond][varname]['contact_max'] = data[varname]
                                                            [xind_contact]
    return results


# the vars of interest
stats_vars = ['HipAnglesX', 'KneeAnglesX', 'AnkleAnglesX',
              'ThoraxAnglesY', 'ThoraxAnglesZ']

# descriptions for the vars
vars_desc = dict()
vars_desc['HipAnglesX'] = 'Hip flexion'
vars_desc['HipAnglesX'] = 'Hip flexion at foot strike'


# data types wanted for each var
wanted = dict()
wanted['HipAnglesX'] = ['at_foot_strike']
wanted['KneeAnglesX'] = ['at_foot_strike']
wanted['AnkleAnglesX'] = ['at_foot_strike', 'contact_max']
wanted['ThoraxAnglesY'] = ['max', 'min']
wanted['ThoraxAnglesZ'] = ['max', 'min']

subjects = ['TD26', 'TD25', 'TD24', 'TD23', 'TD17', 'TD04']
subject = subjects[0]

N_avgdata, N_stddata, C_avgdata, C_stddata = _get_data(subject)

res = _get_vars(N_avgdata, C_avgdata)

# maybe ok but check dorsi/plant vs curves - might be due to no outlier rejection
print('Subject %s' % subject)
for cond in ['N', 'C']:
    for side in ['R', 'L']:
        print('\nCondition: %s' % ('normal' if cond == 'N' else 'cognitive'))
        print('Side: %s' % ('right' if side == 'R' else 'left'))
        print('Hip flexion at foot strike: %.3f' %
              res[cond][side+'HipAnglesX']['at_foot_strike'])
        print('Knee flexion at foot strike: %.3f' %
              res[cond][side+'KneeAnglesX']['at_foot_strike'])
        print('Ankle dorsi/plant at foot strike: %.3f' %
              res[cond][side+'AnkleAnglesX']['at_foot_strike'])
        print('Ankle dorsi/plant max. during contact phase: %.3f' %
              res[cond][side+'AnkleAnglesX']['contact_max'])
        print('Thorax lateral flex max. : %.3f' %
              res[cond][side+'ThoraxAnglesY']['max'])
        print('Thorax lateral flex min. : %.3f' %
              res[cond][side+'ThoraxAnglesY']['min'])
        print('Thorax rotation max. : %.3f' %
              res[cond][side+'ThoraxAnglesZ']['max'])
        print('Thorax rotation min. : %.3f' %
              res[cond][side+'ThoraxAnglesZ']['min'])


""" Write a line that can be pasted into Excel """
row = ''
hdr = ''
print('Subject %s' % subject)
for var in stats_vars:
    for want in wanted[var]:
        for cond in ['N', 'C']:
            for side in ['R', 'L']:
                sidestr = {'R': 'right', 'L': 'left'}
                condstr = {'C': 'cognitive', 'N': 'normal'}
                hdr += '%s_%s (%s %s), ' % (var, want, condstr[cond], sidestr[side])
                row += '%.3f, ' % (res[cond][side+var][want])
print(hdr)
print(row)

from openpyxl import Workbook

wb = Workbook()

dest_filename = 'c:/Temp/test.xlsx'

ws1 = wb.active
ws1.title = "gait data"
ws1.cell(column=1, row=6, value=subject)
hdrli = hdr.split(',')
rowli = row.split(',')
for ind, var in enumerate(hdrli):
    ws1.cell(column=ind+2, row=2, value=var)
    ws1.cell(column=ind+2, row=3, value='degree')    # unit
    ws1.cell(column=ind+2, row=4, value='')    # scale
    ws1.cell(column=ind+2, row=5, value=1)    # type of var = 1 (continuous)
    ws1.cell(column=ind+2, row=6, value=rowli[ind])

wb.save(filename = dest_filename)













